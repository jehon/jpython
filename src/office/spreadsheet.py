
import csv
import pathlib
import re
from typing import cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def worksheet_temp() -> Worksheet:
    """ Create a temp spreadsheet in a new workbook
    """
    return openpyxl.Workbook().active

# pylint: disable-next=too-many-arguments
def spreadsheet_open(
                     filename: str,
                     delimiter: str = ',',
                     csv_title: str = 'csv',
                     read_only: bool = False
                     ) -> openpyxl.Workbook:
    """Load anything as a workbook

            can be xls, xlsx, csv
    """

    match pathlib.Path(filename).suffix:
        case '.csv':
            workbook: openpyxl.Workbook = openpyxl.Workbook()
            worksheet: Worksheet = workbook.active
            worksheet.title = csv_title

            with open(filename, encoding='utf-8') as csv_file:
                reader = csv.reader(csv_file, delimiter=delimiter)
                for row in reader:
                    worksheet.append(row)
            return workbook

        case _:
            return openpyxl.load_workbook(filename=filename, read_only=read_only)

def worksheet_copy(source, target_workbook: openpyxl.Workbook, title: str) -> None:
    # Copying the cell values from template excel file to destination excel file, one by one
    target = target_workbook.create_sheet(title=title)

    for i in range (1, source.max_row  + 1):
        for j in range (1, source.max_column + 1):
            # writing the read value to destination excel file
            target.cell(row = i, column = j, value=source.cell(row = i, column = j).value)

def worksheet_headers(worksheet: Worksheet) -> dict[str, int]:
    """ Get the headers of a worksheet

        Returns a map of headers to column number
    """

    row = worksheet['1']
    headers_map = {}
    for cell in row:
        headers_map[cell.value] = cell.column
    return headers_map

def worksheet_inject(
        source: Worksheet,
        target: Worksheet,
        blanking: int = 50
        ) -> Worksheet :
    """
        Try to map a sheet onto another

        To make this, it will look to target first row, treating it as headers
        For each of these headers, it look in the original spreadsheet to extract the column

        To allow the spreadsheet to have more columns, it stop at the first empty column (or containing 'x')
    """

    target_headers_row = target['1']
    source_headers = worksheet_headers(source)

    mappings = {}

    for cell in target_headers_row:
        if cell.value is None or cell.value == 'x':
            # We stop at the first empty column
            break
        if not cell.value in source_headers:
            raise ValueError(f"Target column '{cell.value}' at {cell.column}"
                + f" is not in the source: {source_headers}")
        mappings[cell.column] = source_headers[cell.value]

    current_row = 0
    for row in source.rows:
        current_row += 1
        row_number = row[0].row
        if row_number == 1:
            # Skip header row
            continue

        for m_src, m_target in mappings.items():
            # for cell in row:
            target.cell(row = current_row, column = m_src, value = row[m_target-1].value)

    for _ in range(blanking):
        current_row += 1
        for m_src, _ in mappings.items():
            target.cell(row = current_row, column = m_src, value = '')

    return target

def inject_worksheet_from_array(data: list[dict], target: Worksheet, blanking = 10) -> Worksheet:
    """
        Try to map a sheet onto another

        To make this, it will look to target first row, treating it as headers
        For each of these headers, it look in the original spreadsheet to extract the column

        To allow the spreadsheet to have more columns, it stop at the first empty column (or containing 'x')
    """
    current_row = 2
    for _, line in enumerate(data):
        for col in range (1, target.max_column + 1):
            header = str(target.cell(row = 1, column = col).value)
            if header is None or header == 'x':
                # We stop at the first empty column
                break
            val = line.get(header)
            if val is None:
                val = ''
            target.cell(row = current_row, column = col, value = val)
        current_row += 1

    for _ in range(blanking):
        current_row += 1
        for col in range (1, target.max_column + 1):
            target.cell(row = current_row, column = col, value = '')

    return target

# pylint: disable-next=too-many-arguments,too-many-locals
def worksheet_split(
        source: Worksheet,
        target: Worksheet,
        key: str,
        value: str,
        split_key: str = "splitted_key",
        split_value: str = "splitted_value",
        regex: str = r"(([A-Z]{3,10}|FM)-?[0-9]{1,4})",
        empty: str = "empty"
        ) -> Worksheet:
    """ Divide each line by looking in dividing_column for regex
        and duplicating it by adding:
         - regex+column
         - target_column
    """
    source_headers = worksheet_headers(source)
    c_key: int = source_headers[key]
    c_value: int = source_headers[value]
    c_split_key: int = len(source_headers) + 1
    c_split_value: int = c_split_key + 1
    compiled_regex: re.Pattern = re.compile(regex)

    nrow = 1
    for row in source.rows:
        # Thanks to https://stackoverflow.com/a/44925460/1954789
        if nrow == 1:
            # Headers
            target.append((cell.value for cell in row))
            target.cell(row=nrow, column=c_split_key, value=split_key)
            target.cell(row=nrow, column=c_split_value, value=split_value)
            nrow += 1
            continue

        if len(row) >= c_key and len(row) >= c_value:
            # Enough data...
            if row[c_key - 1].value:
                matches = compiled_regex.findall(cast(str, row[c_key - 1].value))
                if len(matches) > 0:
                    # Multiple matches
                    hours = float(cast(str, row[c_value - 1].value)) / len(matches)
                    for one_match in matches:
                        target.append((cell.value for cell in row))
                        target.cell(row=nrow, column=c_split_key, value=one_match[0])
                        target.cell(row=nrow, column=c_split_value, value=hours)
                        nrow += 1

                    continue

        # Default case: copy data and add splitted data
        target.append((cell.value for cell in row))
        target.cell(row=nrow, column=c_split_key, value=empty)
        target.cell(row=nrow, column=c_split_value, value=row[c_value - 1].value)
        nrow += 1

    return target

def worksheet_fix_header(worksheet: Worksheet, old: str, new: str) -> None:
    """ Change a header to be another value
    """

    currentHeaders = worksheet_headers(worksheet)
    if currentHeaders.get(new) is None:
        cOld = currentHeaders.get(old)
        if cOld is None:
            raise ValueError(f"Could not find '{old}' in current spreadsheet")

        print(f"Converting '{old}' en '{new}' {cOld}")
        worksheet.cell(row=1, column=cOld, value=new)
